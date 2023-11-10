import openpyxl
import re
import sys
from pathlib import Path

def check_for_duplicates(columns):
    """Check for duplicate column names and return duplicates."""
    seen = set()
    duplicates = set()

    for column in columns:
        if column in seen:
            duplicates.add(column)
        seen.add(column)

    return duplicates

if len(sys.argv) < 2:
    print("Usage: script.py <path_to_xlsx_file>")
    sys.exit(1)

file_name = sys.argv[1]
wb = openpyxl.load_workbook(file_name)
sheet = wb.active

# Identify columns that have a header value (not None or empty string)
valid_columns = [cell.column_letter for cell in sheet[1] if cell.value and str(cell.value).strip() != ""]

# Processing the first row to create column names
columns = [re.sub(r'[^a-zA-Z0-9_]', '', str(cell.value)).replace(' ', '_')[:28].lower() for cell in sheet[1] if cell.column_letter in valid_columns]

# Check for duplicate columns
duplicates = check_for_duplicates(columns)

# Handle duplicate columns (if any)
if duplicates:
    for duplicate in duplicates:
        counter = 1
        while duplicate in columns:
            columns[columns.index(duplicate)] = f"{duplicate}_{counter}"
            counter += 1

# Derive table name from the Excel filename
file_base_name = Path(file_name).stem.replace('.xlsx', '').replace(' ', '_')
table_name = re.sub(r'[^a-zA-Z0-9_]', '', file_base_name).lower()

create_query = f"CREATE TABLE {table_name} (" + ", ".join([f"{column_name} VARCHAR2(4000 CHAR)" for column_name in columns]) + "\n);"

# Determine the index of valid columns
valid_column_indices = [cell.column - 1 for cell in sheet[1] if cell.value and str(cell.value).strip() != ""]

# Processing data rows for INSERT statement
values_list = []

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    valid_data = [row[idx] for idx in valid_column_indices]
    processed_values = [f"'{value}'" if value is not None and value != "" else "NULL" for value in valid_data]
    values_list.append(f"({', '.join(processed_values)})")

insert_queries = f"INSERT INTO {table_name} (" + ", ".join(columns) + ") VALUES" + ",\n".join(values_list) + ";"

# Write CREATE TABLE and INSERT INTO statements to separate txt files
with open('create_table.txt', 'w', encoding='utf-8') as f:
    f.write(create_query)

with open('insert_data.txt', 'w', encoding='utf-8') as f:
    f.write(insert_queries)